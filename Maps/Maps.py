import openpyxl
import requests

api_key = "YOUR API KEY"

# Open the municipalities workbook of target area
municipalities_workbook = openpyxl.load_workbook("Municipalities.xlsx")
results_workbook = openpyxl.Workbook()
results_sheet = results_workbook.active
results_sheet.title = "Your sheet name"

for sheet_name in municipalities_workbook.sheetnames:
    sheet = municipalities_workbook[sheet_name]

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                query = "Query in " + cell.value + ", " + sheet_name + ", Location"
                print("Query:", query)

                
                url = "https://maps.googleapis.com/maps/api/place/textsearch/json?query={}&key={}".format(query, api_key)
                response = requests.get(url)
                data = response.json()

                
                for result in data["results"]:
                    name = result["name"]
                    phone = result.get("formatted_phone_number", "")
                    address = result.get("formatted_address", "")

                   
                    results_sheet.append([name, address, phone])

results_workbook.save("notebookname.xlsx")

